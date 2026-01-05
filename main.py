"""
EduSense AI - Enterprise Backend System
================================================================================
Copyright (c) 2025 EduSense Engineering Team. All Rights Reserved.

PROJECT METADATA:
-----------------
App Name    : EduSense AI Backend
Version     : 6.0.0-Ultimate
Environment : Production / High Availability
License     : Proprietary
Maintainer  : Yasin Muhammad Yusuf

SYSTEM ARCHITECTURE:
--------------------
1.  Core Framework      : FastAPI (ASGI High Performance)
2.  Database ORM        : SQLAlchemy (Asynchronous) + PGVector
3.  Computer Vision     : InsightFace (ArcFace) + OpenCV
4.  Process Management  : Multiprocessing & AsyncIO
5.  Security            : OAuth2 + JWT (RS256/HS256)

MODULES OVERVIEW:
-----------------
[A] System Config       : Logging, Environment, Directories
[B] State Management    : In-Memory Caching (Redis replacement)
[C] Middleware          : CORS, Process Timer, Error Interceptor
[D] Auth & Security     : RBAC, Token Generation, Hash Verification
[E] AI Engine           : Face Detection, Embedding Extraction, Vector Search
[F] Data Controllers    : CRUD for Users, Classes, Schedules
[G] Business Logic      : Attendance Sync, Dispute Handling, Reporting

CHANGELOG (v6.0.0):
-------------------
- [MAJOR] Implemented 'Strict Session Binding' to prevent cross-class attendance.
- [MAJOR] Added 'Smart Upsert' logic for manual attendance to handle disputes.
- [ENHANCE] Added Global Exception Handlers for robust error catching.
- [ENHANCE] Added detailed Docstrings for automated documentation generation.
- [ENHANCE] Added Performance Middleware to track API latency.
================================================================================
"""

import os
import cv2
import uuid
import time
import httpx
import shutil
import logging
import asyncio
import aiofiles
import numpy as np
import pandas as pd
from datetime import datetime, timedelta, date
from typing import List, Optional, Union, Dict, Any, Tuple
from collections import defaultdict
from functools import lru_cache

# --- FASTAPI CORE & UTILITIES ---
from fastapi import (
    FastAPI, 
    UploadFile, 
    File, 
    Form, 
    BackgroundTasks, 
    Depends, 
    HTTPException, 
    status, 
    Request,
    APIRouter,
    Query,
    Path
)
from fastapi.responses import (
    HTMLResponse, 
    FileResponse, 
    RedirectResponse, 
    JSONResponse,
    Response,  # For Excel export
    StreamingResponse  # For Excel streaming
)
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm, OAuth2PasswordBearer
from fastapi.encoders import jsonable_encoder
from fastapi.exceptions import RequestValidationError

# --- AI & COMPUTER VISION LIBRARIES ---
try:
    from insightface.app import FaceAnalysis
except ImportError as e:
    print(f"CRITICAL ERROR: Failed to import InsightFace. {e}")
    # Dummy class to prevent IDE errors during development without GPU
    class FaceAnalysis:
        def __init__(self, **kwargs): pass
        def prepare(self, **kwargs): pass
        def get(self, img): return []

# --- DATABASE DRIVERS & ORM ---
from sqlalchemy.future import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy import desc, update, delete, func, and_, or_
from sqlalchemy.exc import (
    IntegrityError, 
    SQLAlchemyError, 
    NoResultFound, 
    DBAPIError
)

# --- LOCAL APPLICATION MODULES ---
# Mengimpor modul-modul pendukung dari file lain dalam direktori yang sama
from database import engine, Base, get_db, AsyncSessionLocal
from models import (
    Mahasiswa, 
    LogAbsensi, 
    Users, 
    VideoTask, 
    Kelas, 
    Jadwal,
    KelasEnrollment,  # For student class enrollment
    StudentClass  # NEW: Student class groups (A11.4109, etc)
)
from auth_utils import (
    get_password_hash, 
    verify_password, 
    create_access_token, 
    RoleChecker, 
    get_current_user,
    ALGORITHM,
    SECRET_KEY
)
from pydantic import BaseModel, Field, validator, EmailStr

# ==============================================================================
# [SECTION 1] SYSTEM CONFIGURATION & ADVANCED LOGGING
# ==============================================================================

# Setup Direktori Logging
LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

# Konfigurasi Logging Custom dengan Rotasi File (Simulasi)
# Format log mencakup: Timestamp - Level - Module - Message
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s::%(funcName)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(LOG_DIR, "edusense_server.log"), mode='a', encoding='utf-8')
    ]
)
logger = logging.getLogger("EduSenseCore")

# Inisialisasi Aplikasi FastAPI
app = FastAPI(
    title="EduSense AI Backend API",
    description="""
    ## EduSense Enterprise API
    
    Platform Backend untuk manajemen absensi cerdas menggunakan pengenalan wajah.
    
    ### Fitur & Kapabilitas:
    1.  **Sistem Autentikasi**: JWT Token Based Auth dengan Role Management.
    2.  **AI Engine**: InsightFace (ArcFace) untuk ekstraksi fitur wajah.
    3.  **Data Sync**: Sinkronisasi dua arah antara Log Absensi dan Data Master.
    4.  **Reporting**: Export Excel otomatis dengan status kehadiran detail.
    
    ### Kontak Support:
    Tim EduSense - dev@edusense.id
    """,
    version="6.0.0",
    terms_of_service="https://edusense.id/terms",
    contact={
        "name": "EduSense Support Team",
        "url": "https://edusense.id",
        "email": "support@edusense.id",
    },
    license_info={
        "name": "Proprietary Enterprise License",
        "url": "https://edusense.id/license",
    },
    docs_url="/docs",
    redoc_url="/redoc"
)

# Konfigurasi Path Direktori Penyimpanan File Server
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DIRECTORY_CONFIG = {
    "TEMP_VIDEO": os.path.join(BASE_DIR, "temp_videos"),
    "CROP_FACE": os.path.join(BASE_DIR, "hasil_crop"),
    "EXCEL_REPORT": os.path.join(BASE_DIR, "laporan_excel"),
    "STATIC_ASSETS": os.path.join(BASE_DIR, "static"),
}

# Inisialisasi Direktori (Robust Creation)
for key, dir_path in DIRECTORY_CONFIG.items():
    if not os.path.exists(dir_path):
        try:
            os.makedirs(dir_path, exist_ok=True)
            logger.info(f"FileSystem: Created directory at {dir_path}")
        except OSError as e:
            logger.critical(f"FileSystem: Failed to create directory {dir_path}. Error: {e}")
            # Pada sistem produksi, kita mungkin ingin menghentikan server jika direktori kritis gagal dibuat
            # raise e

# Mounting Static Files untuk akses publik via HTTP
app.mount("/hasil_crop", StaticFiles(directory=DIRECTORY_CONFIG["CROP_FACE"]), name="crop_wajah")

# Konfigurasi Middleware CORS (Cross-Origin Resource Sharing)
# Mengizinkan frontend mengakses API ini dari domain yang berbeda
origins = [
    "http://localhost",
    "http://localhost:8000",
    "http://127.0.0.1",
    "http://127.0.0.1:8000",
    "*" # WARNING: Di production, ganti ini dengan domain spesifik untuk keamanan
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
)

# Middleware Kustom: Performance Monitoring
@app.middleware("http")
async def add_process_time_header(request: Request, call_next):
    """
    Middleware untuk mengukur waktu eksekusi setiap request.
    Berguna untuk monitoring latensi server.
    """
    start_time = time.time()
    response = await call_next(request)
    process_time = time.time() - start_time
    response.headers["X-Process-Time"] = str(process_time)
    
    # Log request lambat (> 1 detik)
    if process_time > 1.0:
        logger.warning(f"SLOW REQUEST: {request.url.path} took {process_time:.4f}s")
        
    return response

# Konstanta Konfigurasi AI & Eksternal
API_EMOTION_URL = "https://risetkami-risetkami.hf.space/predict_face"
FACE_SIMILARITY_THRESHOLD = 0.50  # Ambang batas kemiripan wajah (Cosine Similarity)
VIDEO_PROCESSING_FRAME_SKIP = 30  # Skip frame untuk optimasi (Analisis per 1 detik jika FPS=30)

# ==============================================================================
# [SECTION 2] GLOBAL STATE MANAGEMENT (SINGLETON CACHE)
# ==============================================================================

class SystemStateManager:
    """
    Kelas Singleton untuk mengelola state global aplikasi.
    Menyimpan objek berat di memori untuk menghindari reload berulang dan meningkatkan performa.
    """
    def __init__(self):
        # Cache status upload video (Task ID -> Status Dict)
        self.tasks_db: Dict[str, Dict[str, Any]] = {}
        
        # Instance Model InsightFace (Berat, hanya load sekali saat startup)
        self.face_app: Optional[FaceAnalysis] = None 
        
        # List NIM Mahasiswa (Indexing untuk pencarian cepat)
        self.known_ids: List[str] = []
        
        # Matrix Vector Wajah (Numpy Array) untuk kalkulasi jarak cosine super cepat
        self.known_matrix: Optional[np.ndarray] = None 
        
        # Metadata Status Sistem
        self.system_status: str = "STARTING"
        self.start_time: datetime = datetime.now()
        self.last_reload: Optional[datetime] = None
        
        # === CACHE DATA AKADEMIK (Optimasi Query) ===
        self.cache_kelas: Optional[List[Dict]] = None
        self.cache_kelas_time: Optional[datetime] = None
        self.cache_dosen: Optional[List[Dict]] = None
        self.cache_dosen_time: Optional[datetime] = None
        self.cache_jadwal: Optional[List[Dict]] = None
        self.cache_jadwal_time: Optional[datetime] = None
        self.CACHE_TTL_SECONDS: int = 60  # Cache valid selama 60 detik
    
    def invalidate_cache(self, cache_type: str = "all"):
        """Invalidate cache when data is modified."""
        if cache_type in ["all", "kelas"]:
            self.cache_kelas = None
        if cache_type in ["all", "dosen"]:
            self.cache_dosen = None
        if cache_type in ["all", "jadwal"]:
            self.cache_jadwal = None
    
    def is_cache_valid(self, cache_time: Optional[datetime]) -> bool:
        """Check if cache is still valid."""
        if cache_time is None:
            return False
        return (datetime.now() - cache_time).total_seconds() < self.CACHE_TTL_SECONDS

    def update_task(self, task_id: str, status: str, progress: int = 0, error: str = None):
        """
        Helper method untuk update status task video dengan thread-safety minimal.
        """
        payload = {
            "status": status,
            "progress": progress,
            "last_update": datetime.now().isoformat()
        }
        if error:
            payload["error"] = error
            
        # Inisialisasi jika belum ada
        if task_id not in self.tasks_db:
            self.tasks_db[task_id] = {}
            
        self.tasks_db[task_id].update(payload)

# Inisialisasi State Global
state = SystemStateManager()

# ==============================================================================
# [SECTION 3] ROLE BASED ACCESS CONTROL (RBAC) & DEPENDENCIES
# ==============================================================================

# Dependency Injectors untuk membatasi akses endpoint berdasarkan Role User.
# Memastikan user hanya bisa mengakses fitur sesuai hak aksesnya.

allow_kaprodi = RoleChecker(["kaprodi"])
# Dosen memiliki akses ke fitur dosen dan juga fitur umum tertentu
allow_dosen   = RoleChecker(["dosen", "kaprodi"]) 
allow_mhs     = RoleChecker(["mahasiswa"])
# Akses umum untuk semua user yang terautentikasi dalam sistem
allow_all     = RoleChecker(["mahasiswa", "dosen", "kaprodi"])

# ==============================================================================
# [SECTION 4] DATA TRANSFER OBJECTS (DTO) / PYDANTIC SCHEMAS
# ==============================================================================

class UserLoginSchema(BaseModel):
    """Schema untuk validasi input login user."""
    username: str = Field(..., description="Username atau NIM/NPP pengguna", min_length=3)
    password: str = Field(..., description="Password akun")

class DosenCreateSchema(BaseModel):
    """Schema untuk pembuatan akun dosen baru oleh Admin (Kaprodi)."""
    username: str = Field(..., min_length=3, description="Nomor Pokok Pegawai (NPP)")
    password: str = Field(..., min_length=6, description="Password akun minimal 6 karakter")
    full_name: str = Field(..., min_length=3, description="Nama Lengkap beserta Gelar")

    @validator('username')
    def validate_username(cls, v):
        if not v.strip(): 
            raise ValueError("Username tidak boleh kosong atau spasi saja.")
        return v

class DosenUpdateSchema(BaseModel):
    """Schema untuk update data dosen."""
    full_name: str = Field(..., min_length=3, description="Nama Lengkap beserta Gelar")
    password: Optional[str] = Field(None, min_length=6, description="Password baru (opsional)")

class KelasCreateSchema(BaseModel):
    """Schema untuk input data mata kuliah baru."""
    nama_matkul: str = Field(..., description="Nama Mata Kuliah")
    kode_ruang: str = Field(..., description="Kode Ruang (misal: H.4.1)")

class JadwalCreateSchema(BaseModel):
    """Schema untuk pembuatan jadwal kuliah."""
    dosen_username: str = Field(..., description="NPP Dosen Pengampu")
    kelas_id: int = Field(..., description="ID Kelas dari database")
    student_class_id: Optional[int] = Field(None, description="ID Student Class (Kelas Mahasiswa)")
    hari: str = Field(..., description="Hari (Senin-Minggu)")
    jam_mulai: str = Field(..., description="Format HH:MM")
    jam_selesai: str = Field(..., description="Format HH:MM")
    
    @validator('hari')
    def validate_hari(cls, v):
        valid_days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        if v.capitalize() not in valid_days:
            raise ValueError(f"Hari tidak valid. Gunakan salah satu dari: {', '.join(valid_days)}")
        return v.capitalize()

class ManualAbsenSchema(BaseModel):
    """Schema untuk input absen manual oleh Dosen."""
    nim: str = Field(..., description="NIM Mahasiswa yang akan diabsen")
    jadwal_id: int = Field(..., description="ID Jadwal sesi kuliah yang sedang berlangsung")

class DisputeReportSchema(BaseModel):
    """Schema untuk pelaporan ketidaksesuaian absen oleh Mahasiswa."""
    log_id: int = Field(..., description="ID Log Absensi yang disengketakan")
    alasan: str = Field(..., min_length=5, max_length=255, description="Alasan pengajuan laporan")

class UpdatePasswordSchema(BaseModel):
    """Schema untuk update password user (dosen/mahasiswa)."""
    old_password: str = Field(..., min_length=1, description="Password lama")
    new_password: str = Field(..., min_length=6, description="Password baru minimal 6 karakter")
    
    @validator('new_password')
    def validate_new_password(cls, v, values):
        if 'old_password' in values and v == values['old_password']:
            raise ValueError('Password baru harus berbeda dengan password lama')
        return v

# ===== ENROLLMENT SCHEMAS =====
class EnrollmentCreateSchema(BaseModel):
    """Schema untuk mendaftarkan mahasiswa ke kelas mahasiswa (student class)."""
    nim: str = Field(..., description="NIM Mahasiswa yang akan didaftarkan")
    student_class_id: int = Field(..., description="ID Student Class (Kelas Mahasiswa)")

class EnrollmentBulkSchema(BaseModel):
    """Schema untuk bulk enrollment multiple students to one student class."""
    student_class_id: int = Field(..., description="ID Student Class tujuan")
    nim_list: List[str] = Field(..., description="List NIM mahasiswa yang akan didaftarkan")

class EnrollmentBulkSchema(BaseModel):
    """Schema untuk bulk enrollment multiple students to one class."""
    kelas_id: int = Field(..., description="ID Kelas tujuan")
    nim_list: List[str] = Field(..., description="List NIM mahasiswa yang akan didaftarkan")

# ==============================================================================
# [SECTION 5] UTILITY & HELPER FUNCTIONS
# ==============================================================================

def compress_image_to_bytes(image_cv: np.ndarray, max_size: int = 224, quality: int = 60) -> bytes:
    """
    Fungsi utilitas untuk mengompresi gambar OpenCV (Numpy Array) menjadi bytes JPEG.
    Digunakan sebelum mengirim gambar ke API Eksternal untuk menghemat bandwidth network.

    Args:
        image_cv (np.ndarray): Gambar input dalam format BGR.
        max_size (int): Ukuran dimensi maksimum (lebar/tinggi).
        quality (int): Kualitas kompresi JPEG (0-100).

    Returns:
        bytes: Data gambar terkompresi.
    """
    try:
        h, w = image_cv.shape[:2]
        # Resize logic jika gambar terlalu besar
        if max(h, w) > max_size:
            scale = max_size / max(h, w)
            new_w, new_h = int(w * scale), int(h * scale)
            image_cv = cv2.resize(image_cv, (new_w, new_h), interpolation=cv2.INTER_AREA)
        
        # Encoding ke format JPG
        success, encoded_img = cv2.imencode('.jpg', image_cv, [int(cv2.IMWRITE_JPEG_QUALITY), quality])
        if not success:
            logger.warning("Image Utils: Failed to encode image.")
            return b""
        return encoded_img.tobytes()
    except Exception as e:
        logger.error(f"Image Utils: Error compressing image - {e}")
        return b""

def identify_face_fast(embedding: np.ndarray) -> tuple:
    """
    Algoritma pencarian wajah berkinerja tinggi menggunakan Operasi Matriks (Vectorization).
    
    Logic:
    1. Normalisasi vector input.
    2. Hitung Dot Product antara input vector dengan SEMUA vector di database sekaligus.
    3. Ambil nilai maksimum (kemiripan tertinggi).
    4. Bandingkan dengan Threshold.

    Complexity: O(1) effective with Hardware Acceleration (NumPy).

    Returns:
        tuple: (NIM_Found, Similarity_Score) atau ("Unknown", 0.0)
    """
    # Cek ketersediaan database wajah
    if state.known_matrix is None or len(state.known_ids) == 0:
        return "Unknown", 0.0

    try:
        # Normalisasi L2 pada vector input
        norm_emb = embedding / np.linalg.norm(embedding)
        
        # Matrix Multiplication (Dot Product) -> Cosine Similarity
        # Shape: (N_Database, 512) dot (512,) -> (N_Database,)
        scores = np.dot(state.known_matrix, norm_emb)
        
        # Cari index dengan score tertinggi
        best_idx = np.argmax(scores)
        max_score = scores[best_idx]
        
        # Cek apakah score memenuhi ambang batas (Threshold)
        if max_score > FACE_SIMILARITY_THRESHOLD:
            return state.known_ids[best_idx], float(max_score)
        
        return "Unknown", 0.0
    except Exception as e:
        logger.error(f"AI Module: Critical error in face identification - {e}")
        return "Error", 0.0

async def reload_face_database_async():
    """
    Fungsi Async untuk memuat ulang data vektor wajah dari Database PostgreSQL ke RAM.
    Fungsi ini krusial untuk kecepatan sistem. Tanpa ini, sistem harus query DB tiap frame video.
    Dijalankan saat startup dan setelah registrasi mahasiswa baru.
    """
    logger.info("Database: Memulai proses reload database wajah ke RAM...")
    
    temp_ids = []
    temp_vectors = []
    
    try:
        async with AsyncSessionLocal() as db:
            # Mengambil hanya data yang memiliki embedding (wajah terdaftar)
            stmt = select(Mahasiswa.nim, Mahasiswa.embedding_data).where(
                Mahasiswa.embedding_data.isnot(None)
            )
            result = await db.execute(stmt)
            rows = result.all()
            
            for nim, embedding_val in rows:
                try:
                    # Konversi List float -> Numpy Array Float32
                    emb_array = np.array(embedding_val, dtype=np.float32)
                    
                    # Pra-normalisasi vector agar saat runtime tidak perlu hitung ulang norm
                    norm_vec = emb_array / np.linalg.norm(emb_array)
                    
                    temp_ids.append(nim)
                    temp_vectors.append(norm_vec)
                except Exception as inner_e:
                    logger.warning(f"Database: Data wajah korup untuk NIM {nim} - {inner_e}")

        # Update State Global (Atomic Update)
        if temp_vectors:
            state.known_ids = temp_ids
            state.known_matrix = np.array(temp_vectors)
            state.last_reload = datetime.now()
            logger.info(f"✅ Database: Berhasil memuat {len(temp_ids)} profil wajah.")
        else:
            state.known_ids = []
            state.known_matrix = None
            logger.warning("⚠️ Database: Tidak ada data wajah yang ditemukan.")
            
    except Exception as e:
        logger.error(f"❌ Database: Critical Error saat reload - {e}")

# ==============================================================================
# [SECTION 6] APPLICATION LIFECYCLE (STARTUP & SHUTDOWN)
# ==============================================================================

@app.on_event("startup")
async def startup_event_handler():
    """
    Handler yang dijalankan saat aplikasi pertama kali menyala.
    Tugas:
    1. Inisialisasi koneksi database dan skema.
    2. Load Model AI ke GPU/CPU.
    3. Load data wajah ke memori.
    """
    logger.info("🚀 EduSense AI Server Starting Up...")
    state.system_status = "INITIALIZING"
    
    # 1. Database Initialization
    try:
        async with engine.begin() as conn:
            # Create tables if not exist (SQLAlchemy)
            await conn.run_sync(Base.metadata.create_all)
        logger.info("✅ Database: Koneksi berhasil dan skema terverifikasi.")
    except Exception as e:
        logger.critical(f"❌ Database: Koneksi Gagal - {e}")
        # In production, we might want to exit here
        # sys.exit(1)

    # 2. AI Model Initialization
    try:
        logger.info("⏳ AI: Memuat Model InsightFace... (Ini mungkin memakan waktu)")
        # PENTING: Gunakan buffalo_l (sama dengan bulk registration) untuk kompatibilitas embedding
        # buffalo_sc memiliki dimensi embedding berbeda dengan buffalo_l
        providers = ['CUDAExecutionProvider', 'CPUExecutionProvider']
        state.face_app = FaceAnalysis(name='buffalo_l', providers=providers)
        state.face_app.prepare(ctx_id=0, det_size=(640, 640))
        logger.info("✅ AI: Model InsightFace (buffalo_l) Siap.")
    except Exception as e:
        logger.critical(f"❌ AI: Gagal memuat model - {e}")
    
    # 3. Data Preload
    await reload_face_database_async()
    state.system_status = "RUNNING"

@app.on_event("shutdown")
async def shutdown_event_handler():
    """Handler saat aplikasi dimatikan (Cleanup Resources)."""
    logger.info("🛑 EduSense Server Shutting Down...")
    state.system_status = "STOPPED"
    # Close any open connections or files here if needed

# ==============================================================================
# [SECTION 7] SYSTEM HEALTH & DIAGNOSTICS
# ==============================================================================

@app.get("/system/health", tags=["System"])
async def health_check():
    """
    Endpoint publik untuk mengecek kesehatan sistem.
    Digunakan oleh Load Balancer atau Monitoring Tools.
    """
    uptime = datetime.now() - state.start_time
    return {
        "status": state.system_status,
        "uptime": str(uptime),
        "version": app.version,
        "active_tasks": len(state.tasks_db),
        "loaded_faces": len(state.known_ids),
        "last_db_reload": state.last_reload
    }

# ==============================================================================
# [SECTION 8] AUTHENTICATION API ENDPOINTS
# ==============================================================================

@app.post("/token", tags=["Auth"])
async def login_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(), 
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint Login OAuth2 Standard.
    Menerima username & password, memvalidasi, dan mengembalikan Access Token (JWT).
    
    Returns:
        dict: Access Token, Token Type, Role, Username
    """
    logger.info(f"Auth: Login attempt for user {form_data.username}")
    
    try:
        # Query User dari DB
        stmt = select(Users).where(Users.username == form_data.username)
        result = await db.execute(stmt)
        user = result.scalars().first()
        
        # Validasi Kredensial
        if not user or not verify_password(form_data.password, user.password):
            logger.warning(f"Auth: Failed login for {form_data.username} (Invalid Creds)")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Username atau Password salah",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        # Validasi Status Akun
        if not user.is_active:
            logger.warning(f"Auth: Failed login for {form_data.username} (Inactive)")
            raise HTTPException(status_code=400, detail="Akun anda telah dinonaktifkan. Hubungi Admin.")

        # Generate Token
        access_token = create_access_token(data={"sub": user.username, "role": user.role})
        
        logger.info(f"Auth: Success login for {user.username} as {user.role}")
        
        return {
            "access_token": access_token, 
            "token_type": "bearer", 
            "role": user.role, 
            "username": user.full_name
        }
        
    except SQLAlchemyError as e:
        logger.error(f"Auth: DB Error during login - {e}")
        raise HTTPException(status_code=500, detail="Internal Server Error during login")

@app.post("/user/update-password", dependencies=[Depends(allow_all)], tags=["Auth"])
async def update_user_password(
    data: UpdatePasswordSchema,
    user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint untuk update password user (Dosen/Mahasiswa).
    User harus memasukkan password lama dan password baru.
    
    Returns:
        dict: Status update password
    """
    username = user.get('username') or user.get('sub')
    logger.info(f"Auth: Password update attempt for user {username}")
    
    try:
        # Query User dari DB
        stmt = select(Users).where(Users.username == username)
        result = await db.execute(stmt)
        user_obj = result.scalars().first()
        
        if not user_obj:
            logger.error(f"Auth: User {username} not found in database")
            raise HTTPException(status_code=404, detail="User tidak ditemukan")
        
        # Validasi Password Lama
        if not verify_password(data.old_password, user_obj.password):
            logger.warning(f"Auth: Failed password update for {username} (Invalid old password)")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Password lama tidak sesuai"
            )
        
        # Update Password
        user_obj.password = get_password_hash(data.new_password)
        await db.commit()
        
        logger.info(f"Auth: Password successfully updated for {username}")
        
        return {
            "status": "success",
            "message": "Password berhasil diperbarui",
            "username": username
        }
        
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        logger.error(f"Auth: DB Error during password update - {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail="Terjadi kesalahan saat memperbarui password")
    except Exception as e:
        logger.error(f"Auth: Unexpected error during password update - {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail="Terjadi kesalahan internal")

# ==============================================================================
# [SECTION 9] VIDEO PROCESSING ENGINE (BACKGROUND WORKER)
# ==============================================================================

async def background_video_analyzer(task_id: str, file_path: str, dosen_username: str, jadwal_id: int):
    """
    Worker utama untuk memproses video absensi.
    Dijalankan di background thread agar tidak memblokir API Utama.
    
    Workflow:
    1. Buka Video -> Loop Frame
    2. Deteksi Wajah (InsightFace) -> Identifikasi NIM
    3. Simpan Sample Wajah
    4. Analisis Emosi (API Eksternal)
    5. Update Database LogAbsensi dengan logika Strict Binding (Jadwal ID)
    """
    logger.info(f"🎬 [Task {task_id}] Memulai analisis video untuk Jadwal ID: {jadwal_id}")
    state.update_task(task_id, "processing", 0)
    
    async with httpx.AsyncClient() as http_client:
        try:
            # Buka Video Resource
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                raise RuntimeError("Gagal membuka file video. Format mungkin tidak didukung.")

            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            fps = cap.get(cv2.CAP_PROP_FPS) or 30
            
            # Konfigurasi Sampling Frame (Skip frame untuk performa)
            # Analisis 1 frame setiap 1.5 detik
            frame_interval = int(fps * 1.5) 
            current_frame_idx = 0
            
            # Buffer Data Sementara
            # Struktur: { NIM : { 'count': int, 'emosi': { 'happy': 2, ... }, 'sample': str_path } }
            detected_students = defaultdict(lambda: {'count': 0, 'emosi': defaultdict(int), 'sample': None})
            
            # --- VIDEO LOOP ---
            while True:
                # Update Progress ke Cache (untuk polling frontend)
                if total_frames > 0:
                    progress_pct = min(int((current_frame_idx / total_frames) * 100), 99)
                    state.update_task(task_id, "processing", progress_pct)
                
                # Seek & Read Frame
                cap.set(cv2.CAP_PROP_POS_FRAMES, current_frame_idx)
                ret, frame = cap.read()
                
                if not ret: break # End of Video
                
                # --- AI INFERENCE ---
                # Jalankan di thread pool agar loop async tidak terblokir (CPU Bound Operation)
                faces = await asyncio.to_thread(state.face_app.get, frame)
                
                for face in faces:
                    # Identifikasi Wajah (Vector Search)
                    nim_result, score_result = identify_face_fast(face.embedding)
                    
                    if nim_result != "Unknown":
                        # Mahasiswa Teridentifikasi!
                        detected_students[nim_result]['count'] += 1
                        
                        # --- POST-PROCESSING ---
                        # 1. Simpan Sample Foto (Hanya jika belum punya sample terbaik)
                        if detected_students[nim_result]['sample'] is None:
                            b = face.bbox.astype(int)
                            x1, y1 = max(0, b[0]), max(0, b[1])
                            x2, y2 = min(frame.shape[1], b[2]), min(frame.shape[0], b[3])
                            
                            face_crop = frame[y1:y2, x1:x2]
                            
                            if face_crop.size > 0:
                                filename = f"{task_id}_{nim_result}_{current_frame_idx}.jpg"
                                save_fullpath = os.path.join(DIRECTORY_CONFIG["CROP_FACE"], filename)
                                cv2.imwrite(save_fullpath, face_crop)
                                
                                detected_students[nim_result]['sample'] = f"/hasil_crop/{filename}"

                                # 2. Analisis Emosi (Opsional & Fault Tolerant)
                                try:
                                    img_payload = compress_image_to_bytes(face_crop)
                                    response = await http_client.post(
                                        API_EMOTION_URL, 
                                        files={'file': (filename, img_payload, 'image/jpeg')}, 
                                        timeout=2.0 # Fast timeout
                                    )
                                    if response.status_code == 200:
                                        data = response.json()
                                        emotion = data.get('predicted', 'neutral')
                                        detected_students[nim_result]['emosi'][emotion] += 1
                                except Exception:
                                    # Fail silently untuk emosi, jangan hentikan proses utama
                                    pass

                # Lompat ke frame berikutnya
                current_frame_idx += frame_interval
                if current_frame_idx >= total_frames: break
            
            cap.release()
            
            # --- DATABASE UPDATE (TRANSACTIONAL) ---
            logger.info(f"💾 [Task {task_id}] Menyimpan hasil analisis ke Database...")
            logger.info(f"[Task {task_id}] Total mahasiswa terdeteksi di video: {len(detected_students)}")
            
            async with AsyncSessionLocal() as db_session:
                # 1. Update Task Status -> Completed
                task_query = await db_session.execute(select(VideoTask).where(VideoTask.task_id == task_id))
                task_obj = task_query.scalars().first()
                if task_obj:
                    task_obj.status = "completed"
                
                # 2. Get student_class_id dari Jadwal untuk filtering enrollment
                jadwal_query = await db_session.execute(select(Jadwal).where(Jadwal.jadwal_id == jadwal_id))
                jadwal_obj = jadwal_query.scalars().first()
                
                if not jadwal_obj:
                    logger.warning(f"[Task {task_id}] Jadwal tidak ditemukan")
                    state.update_task(task_id, "failed", error="Jadwal tidak ditemukan")
                    return
                
                # 3. Get list of enrolled students untuk validasi (HANYA jika student_class_id ada)
                enrolled_nims = None
                if jadwal_obj.student_class_id:
                    enrolled_query = await db_session.execute(
                        select(KelasEnrollment.nim).where(
                            KelasEnrollment.student_class_id == jadwal_obj.student_class_id
                        )
                    )
                    enrolled_nims = set([row[0] for row in enrolled_query.fetchall()])
                    logger.info(f"[Task {task_id}] Mahasiswa enrolled di kelas ini: {len(enrolled_nims)}")
                else:
                    logger.warning(f"[Task {task_id}] Jadwal tidak memiliki student_class_id, skip enrollment filtering")
                
                # 4. Proses Log Absensi (Smart Upsert with Strict Session + Enrollment Check)
                today_date = datetime.now().date()
                skipped_count = 0
                processed_count = 0
                
                for nim, meta in detected_students.items():
                    # Validasi enrollment: skip jika enrollment filtering aktif DAN mahasiswa tidak enrolled
                    if enrolled_nims is not None and nim not in enrolled_nims:
                        logger.info(f"[Task {task_id}] Skipping NIM {nim} - tidak terdaftar di kelas ini")
                        skipped_count += 1
                        continue
                    # Tentukan emosi dominan
                    dominant_emotion = "Neutral"
                    if meta['emosi']:
                        dominant_emotion = max(meta['emosi'], key=meta['emosi'].get)
                    
                    # Cek Data Existing Hari Ini PADA JADWAL INI
                    # Ini mencegah 'Pergaulan Bebas' Data (Absen lintas kelas)
                    log_stmt = select(LogAbsensi).where(
                        and_(
                            LogAbsensi.nim == nim, 
                            LogAbsensi.jadwal_id == jadwal_id, # Strict Filter
                            func.date(LogAbsensi.waktu_absen) == today_date
                        )
                    )
                    existing_log = (await db_session.execute(log_stmt)).scalars().first()
                    
                    if not existing_log:
                        # CASE A: Belum absen di kelas ini -> Insert Baru
                        new_log = LogAbsensi(
                            task_id=task_id,
                            nim=nim,
                            jadwal_id=jadwal_id, # KUNCI PENTING
                            waktu_absen=func.now(),
                            metode="AI_VIDEO",
                            jumlah_muncul=meta['count'],
                            emosi_dominan=dominant_emotion,
                            bukti_foto=meta['sample'],
                            is_disputed=False
                        )
                        db_session.add(new_log)
                        processed_count += 1
                        logger.info(f"[Task {task_id}] Insert log baru untuk NIM {nim}")
                    else:
                        # CASE B: Sudah absen -> Update Statistik
                        # Hanya update jika metode sebelumnya juga AI
                        if existing_log.metode == "AI_VIDEO":
                            existing_log.jumlah_muncul += meta['count']
                            processed_count += 1
                            logger.info(f"[Task {task_id}] Update log existing untuk NIM {nim}")
                            # Bisa update emosi jika perlu
                            # existing_log.bukti_foto = meta['sample'] 
                
                await db_session.commit()
                logger.info(f"[Task {task_id}] Database commit successful. Total processed: {processed_count}")

            # Finalisasi State
            state.update_task(task_id, "completed", 100)
            enrolled_detected = processed_count
            logger.info(f"✅ [Task {task_id}] Selesai. {enrolled_detected} mahasiswa berhasil diproses (skipped: {skipped_count}).")

        except Exception as e:
            logger.error(f"❌ [Task {task_id}] Error Processing Video: {e}")
            state.update_task(task_id, "failed", error=str(e))
            
            # Mark as Failed in DB
            async with AsyncSessionLocal() as db_err:
                task_q = await db_err.execute(select(VideoTask).where(VideoTask.task_id == task_id))
                t_obj = task_q.scalars().first()
                if t_obj:
                    t_obj.status = "failed"
                    await db_err.commit()

        finally:
            # Cleanup File Sementara
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    logger.info(f"🧹 [Task {task_id}] File temp dihapus.")
                except Exception: pass

@app.post("/analyze/", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def endpoint_analyze_video(
    bg: BackgroundTasks, 
    file: UploadFile = File(...), 
    jadwal_id: int = Form(...), # Parameter Wajib: Jadwal ID
    user: dict = Depends(get_current_user), 
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint untuk Dosen mengupload video kelas dan memicu analisis AI.
    Video akan diproses secara asynchronous (Background Task).
    """
    if state.face_app is None:
        raise HTTPException(
            status_code=503, 
            detail="Model AI sedang loading atau tidak tersedia. Coba sesaat lagi."
        )
    
    # Validasi Kepemilikan Jadwal (Security)
    jadwal_stmt = select(Jadwal).where(Jadwal.jadwal_id == jadwal_id)
    jadwal = (await db.execute(jadwal_stmt)).scalars().first()
    
    if not jadwal:
        raise HTTPException(404, "Jadwal tidak ditemukan")
    
    # Cek apakah dosen yang login adalah pemilik jadwal
    username_login = user.get('username') or user.get('sub')
    if jadwal.dosen_username != username_login:
        raise HTTPException(403, "Anda tidak memiliki izin untuk mengupload video ke jadwal ini.")

    # Validasi Tipe File
    if not file.content_type.startswith('video/'):
        raise HTTPException(400, "File harus berupa video (MP4/MKV/AVI)")

    # Generate Task ID
    task_id = str(uuid.uuid4())
    temp_file_path = os.path.join(DIRECTORY_CONFIG["TEMP_VIDEO"], f"{task_id}.mp4")
    
    # Simpan File (Async I/O)
    try:
        async with aiofiles.open(temp_file_path, 'wb') as out_file:
            while chunk := await file.read(1024 * 1024): # 1MB Chunks
                await out_file.write(chunk)
    except Exception as e:
        logger.error(f"File upload error: {e}")
        raise HTTPException(500, "Gagal menyimpan file video ke server.")

    # Simpan Task Metadata ke DB
    new_task = VideoTask(
        task_id=task_id, 
        dosen_username=username_login, 
        jadwal_id=jadwal_id, # Bind to Jadwal
        filename=file.filename, 
        status="processing"
    )
    db.add(new_task)
    await db.commit()
    
    # Inisialisasi Status di Memory Cache
    state.update_task(task_id, "queued", 0)
    
    # Trigger Background Worker
    bg.add_task(background_video_analyzer, task_id, temp_file_path, username_login, jadwal_id)
    
    return {
        "task_id": task_id, 
        "message": "Video berhasil diunggah. Analisis berjalan di background.",
        "status": "queued"
    }

@app.get("/status/{task_id}", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def endpoint_get_task_status(task_id: str):
    """
    Long-polling endpoint untuk mengecek status analisis video.
    """
    task_info = state.tasks_db.get(task_id)
    if not task_info:
        return {"status": "not_found", "progress": 0}
    return task_info

# ==============================================================================
# [SECTION 10] DOSEN & CLASS MANAGEMENT (STRICT SESSION LOGIC)
# ==============================================================================

@app.get("/dosen/class/{jadwal_id}/attendance", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def get_session_attendance_sync(
    jadwal_id: int,
    db: AsyncSession = Depends(get_db)
):
    """
    [SYNC ENDPOINT] Endpoint Kritis untuk Dashboard Dosen.
    Menggabungkan Data Master Mahasiswa dengan Data Log Absensi pada JADWAL TERTENTU.
    
    Logic:
    1. Ambil student_class_id dari jadwal
    2. Ambil HANYA mahasiswa yang enrolled di student_class tersebut
    3. Ambil log absensi HANYA untuk jadwal_id tersebut pada hari ini
    4. Join data di Python
    """
    try:
        # 1. Get jadwal info to find student_class_id
        jadwal_stmt = select(Jadwal).where(Jadwal.jadwal_id == jadwal_id)
        jadwal_result = await db.execute(jadwal_stmt)
        jadwal = jadwal_result.scalars().first()
        
        if not jadwal:
            raise HTTPException(404, "Jadwal tidak ditemukan")
        
        if not jadwal.student_class_id:
            raise HTTPException(400, "Jadwal tidak memiliki kelas mahasiswa terdaftar")
        
        # 2. Fetch HANYA Mahasiswa yang ENROLLED di student_class ini + Nama User
        stmt_students = select(Mahasiswa).options(joinedload(Mahasiswa.user)).join(
            KelasEnrollment,
            and_(
                KelasEnrollment.nim == Mahasiswa.nim,
                KelasEnrollment.student_class_id == jadwal.student_class_id
            )
        ).where(
            Mahasiswa.embedding_data.isnot(None)  # Hanya yang sudah registrasi wajah
        )
        result_students = await db.execute(stmt_students)
        enrolled_students = result_students.scalars().all()

        # 3. Fetch Log Absensi HARI INI & JADWAL INI
        today = datetime.now().date()
        stmt_logs = select(LogAbsensi).where(
            and_(
                LogAbsensi.jadwal_id == jadwal_id, # Strict Filtering
                func.date(LogAbsensi.waktu_absen) == today
            )
        )
        result_logs = await db.execute(stmt_logs)
        logs_list = result_logs.scalars().all()
        
        # 4. Hash Map untuk lookup cepat
        attendance_map = {log.nim: log for log in logs_list}

        # 4. Hash Map untuk lookup cepat
        attendance_map = {log.nim: log for log in logs_list}

        # 5. Data Merging Logic - HANYA untuk enrolled students
        final_response_data = []
        
        for student in enrolled_students:
            log_data = attendance_map.get(student.nim)
            
            # Default Values
            status = "ALPHA"
            metode = "-"
            is_disputed = False
            
            if log_data:
                metode = log_data.metode
                if log_data.is_disputed:
                    status = "LAPOR" # Menandakan sengketa
                    is_disputed = True
                elif log_data.metode == "ALPHA_SYSTEM":
                    status = "ALPHA"  # Log ALPHA dari tutup kelas
                else:
                    status = "HADIR"
            
            # Nama Handling (Fallback ke NIM jika User belum dilink)
            full_name = student.user.full_name if student.user else f"Mahasiswa {student.nim}"

            final_response_data.append({
                "nim": student.nim,
                "nama": full_name,
                "status": status,
                "metode": metode,
                "is_disputed": is_disputed,
                # [NEW] Data laporan untuk modal dosen
                "keterangan_report": log_data.keterangan_report if log_data else None,
                "bukti_foto": log_data.bukti_foto if log_data else None
            })
            
        # 6. Sorting (Priority: Lapor > Hadir > Alpha, lalu by Nama)
        final_response_data.sort(key=lambda x: (
            x['status'] != 'LAPOR', 
            x['status'] != 'HADIR', 
            x['nama']
        ))
        
        return final_response_data

    except SQLAlchemyError as e:
        logger.error(f"Database error in sync attendance: {e}")
        raise HTTPException(500, "Gagal mengambil data sinkronisasi kelas.")

@app.post("/dosen/manual_absen", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def endpoint_manual_absen(data: ManualAbsenSchema, db: AsyncSession = Depends(get_db)):
    """
    [SMART UPSERT] Menangani Absen Manual & Verifikasi Dispute.
    Sekarang terikat pada JADWAL ID.
    """
    # 1. Validasi Keberadaan Mahasiswa
    check_stmt = select(Mahasiswa).where(Mahasiswa.nim == data.nim)
    if not (await db.execute(check_stmt)).scalars().first():
        raise HTTPException(404, f"NIM {data.nim} tidak ditemukan.")
    
    # 2. Validasi Enrollment: pastikan mahasiswa terdaftar di kelas ini
    jadwal_stmt = select(Jadwal).where(Jadwal.jadwal_id == data.jadwal_id)
    jadwal = (await db.execute(jadwal_stmt)).scalars().first()
    if not jadwal:
        raise HTTPException(404, "Jadwal tidak ditemukan.")
    
    enrollment_stmt = select(KelasEnrollment).where(
        and_(
            KelasEnrollment.nim == data.nim,
            KelasEnrollment.student_class_id == jadwal.student_class_id
        )
    )
    enrollment = (await db.execute(enrollment_stmt)).scalars().first()
    if not enrollment:
        raise HTTPException(403, f"Mahasiswa dengan NIM {data.nim} tidak terdaftar di kelas ini.")
    
    # 3. Cek Log Hari Ini pada Jadwal Ini
    today = datetime.now().date()
    log_stmt = select(LogAbsensi).where(
        and_(
            LogAbsensi.nim == data.nim, 
            LogAbsensi.jadwal_id == data.jadwal_id, # Strict Filter
            func.date(LogAbsensi.waktu_absen) == today
        )
    )
    existing_log = (await db.execute(log_stmt)).scalars().first()
    
    msg = ""
    
    if existing_log:
        # Case Update: Verifikasi Laporan atau Koreksi Manual
        existing_log.metode = "MANUAL_DOSEN"
        existing_log.is_disputed = False  # Clear Sengketa
        existing_log.keterangan_report = None
        existing_log.waktu_absen = func.now()
        msg = f"Status presensi {data.nim} di kelas ini DIPERBARUI menjadi HADIR."
    else:
        # Case Insert: Data Baru
        new_log = LogAbsensi(
            nim=data.nim,
            jadwal_id=data.jadwal_id, # Bind to Schedule
            metode="MANUAL_DOSEN",
            jumlah_muncul=1,
            emosi_dominan="Manual Input",
            waktu_absen=func.now(),
            is_disputed=False
        )
        db.add(new_log)
        msg = f"Mahasiswa {data.nim} DIABSEN manual di kelas ini."
        
    await db.commit()
    return {"status": "success", "message": msg}

@app.post("/dosen/close_class", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def close_class_session(
    jadwal_id: int = Form(...),
    db: AsyncSession = Depends(get_db)
):
    """
    [TUTUP KELAS] Menandai sesi kelas selesai dan membuat log ALPHA untuk semua 
    mahasiswa yang ENROLLED di kelas ini tetapi belum tercatat kehadirannya.
    
    [UPDATED] Hanya mahasiswa yang terdaftar di kelas ini yang akan mendapat ALPHA.
    """
    logger.info(f"Dosen: Menutup sesi kelas untuk jadwal_id={jadwal_id}")
    
    today = datetime.now().date()
    
    try:
        # 1. Get kelas_id from jadwal
        jadwal_stmt = select(Jadwal).where(Jadwal.jadwal_id == jadwal_id)
        jadwal = (await db.execute(jadwal_stmt)).scalars().first()
        if not jadwal:
            raise HTTPException(404, "Jadwal tidak ditemukan")
        
        student_class_id = jadwal.student_class_id  # Changed from kelas_id
        
        # 2. [UPDATED] Ambil mahasiswa yang ENROLLED di student class ini DAN punya data wajah
        enrolled_stmt = select(Mahasiswa.nim).join(
            KelasEnrollment,
            and_(
                KelasEnrollment.nim == Mahasiswa.nim,
                KelasEnrollment.student_class_id == student_class_id  # Changed from kelas_id
            )
        ).where(
            Mahasiswa.embedding_data.isnot(None)
        )
        enrolled_results = (await db.execute(enrolled_stmt)).scalars().all()
        enrolled_nims = set(enrolled_results)
        
        # 3. Ambil mahasiswa yang sudah absen hari ini di jadwal ini
        present_stmt = select(LogAbsensi.nim).where(
            and_(
                LogAbsensi.jadwal_id == jadwal_id,
                func.date(LogAbsensi.waktu_absen) == today
            )
        )
        present_results = (await db.execute(present_stmt)).scalars().all()
        present_nims = set(present_results)
        
        # 4. Buat log ALPHA untuk yang enrolled tapi belum absen
        alpha_count = 0
        for nim in enrolled_nims:
            if nim not in present_nims:
                # Buat log ALPHA
                alpha_log = LogAbsensi(
                    nim=nim,
                    jadwal_id=jadwal_id,
                    metode="ALPHA_SYSTEM",  # Tandai sebagai ALPHA otomatis
                    jumlah_muncul=0,
                    emosi_dominan=None,
                    waktu_absen=func.now(),
                    is_disputed=False
                )
                db.add(alpha_log)
                alpha_count += 1
        
        # 5. Tandai semua VideoTask dari jadwal ini sebagai closed
        await db.execute(
            update(VideoTask)
            .where(VideoTask.jadwal_id == jadwal_id)
            .values(is_closed=True)
        )
        
        await db.commit()
        
        logger.info(f"Dosen: Sesi jadwal_id={jadwal_id} ditutup. {alpha_count} mahasiswa ditandai ALPHA dari {len(enrolled_nims)} enrolled.")
        
        return {
            "status": "success",
            "message": f"Kelas berhasil ditutup. {len(present_nims)} hadir, {alpha_count} alpha.",
            "hadir": len(present_nims),
            "alpha": alpha_count
        }
        
    except SQLAlchemyError as e:
        logger.error(f"Database error closing class: {e}")
        raise HTTPException(500, "Gagal menutup kelas.")


@app.get("/dosen/class_report/{jadwal_id}", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def generate_class_attendance_report(jadwal_id: int, db: AsyncSession = Depends(get_db)):
    """
    Generate Excel report untuk kelas (semua mahasiswa: HADIR & ALPHA).
    Includes keterangan waktu download laporan.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    today = datetime.now().date()
    report_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    try:
        # 1. Ambil info jadwal
        jadwal_stmt = select(Jadwal, Kelas.nama_matkul, Kelas.kode_ruang)\
            .join(Kelas, Jadwal.kelas_id == Kelas.kelas_id)\
            .where(Jadwal.jadwal_id == jadwal_id)
        jadwal_result = (await db.execute(jadwal_stmt)).first()
        
        if not jadwal_result:
            raise HTTPException(404, "Jadwal tidak ditemukan")
        
        jadwal, nama_matkul, ruang = jadwal_result
        
        # 2. Ambil semua mahasiswa yang terdaftar di kelas ini (enrolled + registered face)
        if not jadwal.student_class_id:
            raise HTTPException(400, "Jadwal tidak memiliki student_class_id")
        
        all_mhs_stmt = select(Mahasiswa, Users.full_name)\
            .outerjoin(Users, Mahasiswa.nim == Users.username)\
            .join(
                KelasEnrollment,
                and_(
                    KelasEnrollment.nim == Mahasiswa.nim,
                    KelasEnrollment.student_class_id == jadwal.student_class_id
                )
            )\
            .where(Mahasiswa.embedding_data.isnot(None))
        mhs_results = (await db.execute(all_mhs_stmt)).all()
        
        # 3. Ambil log absensi hari ini
        logs_stmt = select(LogAbsensi).where(
            and_(
                LogAbsensi.jadwal_id == jadwal_id,
                func.date(LogAbsensi.waktu_absen) == today
            )
        )
        logs = (await db.execute(logs_stmt)).scalars().all()
        log_map = {log.nim: log for log in logs}
        
        # 4. Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Kehadiran"
        
        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        table_header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        table_header_font = Font(bold=True, color="FFFFFF")
        hadir_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        alpha_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.merge_cells('A1:E1')
        ws['A1'] = f"LAPORAN KEHADIRAN - {nama_matkul}"
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Ruang: {ruang} | Hari: {jadwal.hari} | Jam: {jadwal.jam_mulai}-{jadwal.jam_selesai}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f"Laporan Digenerate: {report_time}"
        ws['A3'].font = Font(italic=True, size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Table Headers
        headers = ["No", "NIM", "Nama Mahasiswa", "Status", "Metode/Waktu"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data
        hadir_count = 0
        alpha_count = 0
        
        for idx, (mhs, full_name) in enumerate(mhs_results, 1):
            row = 5 + idx
            log = log_map.get(mhs.nim)
            
            # Determine status
            if log:
                if log.metode == "ALPHA_SYSTEM":
                    status = "ALPHA"
                    metode_info = "Tidak Hadir"
                    fill = alpha_fill
                    alpha_count += 1
                else:
                    status = "HADIR"
                    waktu = log.waktu_absen.strftime("%H:%M") if log.waktu_absen else "-"
                    metode_info = f"{log.metode} ({waktu})"
                    fill = hadir_fill
                    hadir_count += 1
            else:
                status = "ALPHA"
                metode_info = "Tidak tercatat"
                fill = alpha_fill
                alpha_count += 1
            
            nama = full_name or f"Mahasiswa {mhs.nim}"
            
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=mhs.nim).border = thin_border
            ws.cell(row=row, column=3, value=nama).border = thin_border
            
            status_cell = ws.cell(row=row, column=4, value=status)
            status_cell.fill = fill
            status_cell.border = thin_border
            status_cell.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=5, value=metode_info).border = thin_border
        
        # Summary
        summary_row = 5 + len(mhs_results) + 2
        ws.cell(row=summary_row, column=1, value="TOTAL:")
        ws.cell(row=summary_row, column=2, value=f"HADIR: {hadir_count}")
        ws.cell(row=summary_row, column=3, value=f"ALPHA: {alpha_count}")
        ws.cell(row=summary_row, column=4, value=f"TOTAL: {len(mhs_results)}")
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Laporan_{nama_matkul.replace(' ', '_')}_{today.strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except SQLAlchemyError as e:
        logger.error(f"Database error generating report: {e}")
        raise HTTPException(500, "Gagal membuat laporan.")



@app.get("/history/tasks", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def get_upload_history(
    user: dict = Depends(get_current_user), 
    db: AsyncSession = Depends(get_db)
):
    """
    Mendapatkan riwayat upload video yang kelasnya sudah ditutup.
    Dosen hanya melihat miliknya sendiri (Row-Level Security).
    [UPDATED] Hanya menampilkan video dari kelas yang sudah ditutup (is_closed=True)
    """
    # Join dengan Jadwal, Kelas, dan Users (untuk nama dosen)
    stmt = select(VideoTask, Jadwal, Kelas, Users.full_name)\
        .join(Jadwal, VideoTask.jadwal_id == Jadwal.jadwal_id, isouter=True)\
        .join(Kelas, Jadwal.kelas_id == Kelas.kelas_id, isouter=True)\
        .join(Users, VideoTask.dosen_username == Users.username, isouter=True)\
        .where(VideoTask.is_closed == True)\
        .order_by(desc(VideoTask.created_at)).limit(50)
    
    # Filter by Role (Security)
    if user['role'] == 'dosen':
        username = user.get('username') or user.get('sub')
        stmt = stmt.where(VideoTask.dosen_username == username)
        
    result = await db.execute(stmt)
    
    output = []
    for task, jad, kel, dosen_name in result.all():
        matkul_name = "Unknown"
        if kel:
            matkul_name = f"{kel.nama_matkul} ({kel.kode_ruang})"
            
        output.append({
            "task_id": task.task_id,
            "filename": task.filename,
            "status": task.status,
            "created_at": task.created_at,
            "matkul": matkul_name,
            "dosen_username": task.dosen_username,
            "dosen_name": dosen_name or task.dosen_username  # Fallback ke username jika nama tidak ada
        })
        
    return output

# ==============================================================================
# [SECTION 11] KAPRODI & ADMIN MANAGEMENT
# ==============================================================================

@app.get("/admin/jadwal", dependencies=[Depends(allow_dosen)], tags=["Admin"])
async def get_all_schedules(
    user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Mengambil data jadwal kuliah.
    - KAPRODI: Melihat Semua.
    - DOSEN: Melihat Milik Sendiri.
    """
    stmt = select(
        Jadwal, 
        Users.full_name, 
        Kelas.nama_matkul, 
        Kelas.kode_ruang,
        StudentClass.class_name
    ).join(Users, Jadwal.dosen_username == Users.username)\
     .join(Kelas, Jadwal.kelas_id == Kelas.kelas_id)\
     .outerjoin(StudentClass, Jadwal.student_class_id == StudentClass.class_id)
    
    # Row-Level Security for Dosen
    if user['role'] == 'dosen':
        username = user.get('username') or user.get('sub')
        stmt = stmt.where(Jadwal.dosen_username == username)
    
    result = await db.execute(stmt)
    
    data = []
    for row in result.all():
        data.append({
            "jadwal_id": row[0].jadwal_id,
            "hari": row[0].hari,
            "jam": f"{row[0].jam_mulai}-{row[0].jam_selesai}",
            "jam_mulai": row[0].jam_mulai,
            "jam_selesai": row[0].jam_selesai,
            "dosen": row[1],
            "dosen_username": row[0].dosen_username,
            "matkul": f"{row[2]} ({row[3]})",
            "kelas_id": row[0].kelas_id,
            "student_class_id": row[0].student_class_id,
            "student_class_name": row[4]
        })
    return data

@app.post("/admin/dosen", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_create_dosen(data: DosenCreateSchema, db: AsyncSession = Depends(get_db)):
    dup = await db.execute(select(Users).where(Users.username == data.username))
    if dup.scalars().first(): raise HTTPException(400, "Username/NPP sudah ada.")
    
    new_user = Users(
        username=data.username,
        password=get_password_hash(data.password),
        full_name=data.full_name,
        role="dosen"
    )
    db.add(new_user)
    await db.commit()
    state.invalidate_cache("dosen")  # Invalidate cache after write
    return {"msg": "Akun Dosen berhasil dibuat"}

@app.get("/admin/dosen", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_list_dosen(db: AsyncSession = Depends(get_db)):
    """List dosen dengan in-memory caching."""
    if state.is_cache_valid(state.cache_dosen_time) and state.cache_dosen:
        return state.cache_dosen
    
    res = await db.execute(select(Users).where(Users.role == "dosen"))
    dosen_list = res.scalars().all()
    state.cache_dosen = [{"username": d.username, "full_name": d.full_name, "is_active": d.is_active} for d in dosen_list]
    state.cache_dosen_time = datetime.now()
    return state.cache_dosen

@app.delete("/admin/dosen/{username}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_delete_dosen(username: str, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(Users).where(Users.username == username))
    await db.commit()
    state.invalidate_cache("dosen")  # Invalidate cache
    return {"msg": "Akun Dosen dihapus"}

@app.put("/admin/dosen/{username}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_update_dosen(username: str, data: DosenUpdateSchema, db: AsyncSession = Depends(get_db)):
    """Update data dosen (nama dan/atau password)."""
    # Cari dosen
    stmt = select(Users).where(Users.username == username, Users.role == "dosen")
    result = await db.execute(stmt)
    dosen = result.scalars().first()
    
    if not dosen:
        raise HTTPException(404, "Dosen tidak ditemukan")
    
    # Update nama
    dosen.full_name = data.full_name
    
    # Update password jika diberikan
    if data.password:
        dosen.password = get_password_hash(data.password)
    
    await db.commit()
    state.invalidate_cache("dosen")  # Invalidate cache
    
    logger.info(f"Admin: Updated dosen {username} - {data.full_name}")
    return {"msg": "Data Dosen berhasil diperbarui"}

@app.get("/admin/kelas", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_list_kelas(db: AsyncSession = Depends(get_db)):
    """List kelas dengan in-memory caching untuk performa optimal."""
    # Check cache first
    if state.is_cache_valid(state.cache_kelas_time) and state.cache_kelas:
        logger.debug("Cache HIT: /admin/kelas")
        return state.cache_kelas
    
    # Cache miss - query database
    logger.debug("Cache MISS: /admin/kelas - querying DB")
    res = await db.execute(select(Kelas))
    kelas_list = res.scalars().all()
    
    # Convert to dict for caching (ORM objects can't be cached directly)
    state.cache_kelas = [{"kelas_id": k.kelas_id, "nama_matkul": k.nama_matkul, "kode_ruang": k.kode_ruang} for k in kelas_list]
    state.cache_kelas_time = datetime.now()
    
    return state.cache_kelas

@app.post("/admin/kelas", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_create_kelas(data: KelasCreateSchema, db: AsyncSession = Depends(get_db)):
    db.add(Kelas(nama_matkul=data.nama_matkul, kode_ruang=data.kode_ruang))
    await db.commit()
    state.invalidate_cache("kelas")  # Invalidate cache after write
    return {"msg": "Kelas berhasil dibuat"}

@app.post("/admin/jadwal", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_create_jadwal(data: JadwalCreateSchema, db: AsyncSession = Depends(get_db)):
    db.add(Jadwal(
        dosen_username=data.dosen_username, 
        kelas_id=data.kelas_id, 
        student_class_id=data.student_class_id,
        hari=data.hari, 
        jam_mulai=data.jam_mulai, 
        jam_selesai=data.jam_selesai
    ))
    await db.commit()
    state.invalidate_cache("jadwal")  # Invalidate cache after write
    return {"msg": "Jadwal berhasil ditambahkan"}

@app.delete("/admin/mahasiswa/{nim}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_delete_mhs(nim: str, db: AsyncSession = Depends(get_db)):
    await db.execute(delete(Mahasiswa).where(Mahasiswa.nim == nim))
    await db.commit()
    await reload_face_database_async()
    return {"msg": "Data Mahasiswa dihapus"}

# Schema untuk create mahasiswa
class MahasiswaCreateSchema(BaseModel):
    nim: str = Field(..., min_length=5, description="NIM Mahasiswa")
    full_name: str = Field(..., min_length=3, description="Nama Lengkap Mahasiswa")
    password: str = Field(default="password123", description="Password default untuk login")

@app.post("/admin/mahasiswa", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def admin_create_mahasiswa(data: MahasiswaCreateSchema, db: AsyncSession = Depends(get_db)):
    """
    Tambah mahasiswa baru (tanpa foto wajah).
    Mahasiswa perlu registrasi wajah untuk bisa absen AI.
    """
    # Check duplicate
    dup_mhs = await db.execute(select(Mahasiswa).where(Mahasiswa.nim == data.nim))
    if dup_mhs.scalars().first():
        raise HTTPException(400, f"NIM {data.nim} sudah terdaftar")
    
    dup_user = await db.execute(select(Users).where(Users.username == data.nim))
    if dup_user.scalars().first():
        raise HTTPException(400, f"Username {data.nim} sudah ada di sistem")
    
    # Create User account
    new_user = Users(
        username=data.nim,
        password=get_password_hash(data.password),
        full_name=data.full_name,
        role="mahasiswa",
        is_active=True
    )
    db.add(new_user)
    
    # Create Mahasiswa record (without face embedding)
    new_mhs = Mahasiswa(
        nim=data.nim,
        embedding_data=None  # Will be added later via face registration
    )
    db.add(new_mhs)
    
    await db.commit()
    
    logger.info(f"Admin: Created new student {data.nim} - {data.full_name}")
    return {
        "msg": f"Mahasiswa {data.full_name} berhasil ditambahkan",
        "nim": data.nim,
        "note": "Mahasiswa perlu registrasi wajah untuk menggunakan fitur absensi AI"
    }

# ==============================================================================
# [SECTION 11.5] ENROLLMENT MANAGEMENT (Pendaftaran Mahasiswa ke Kelas)
# ==============================================================================

@app.post("/admin/enrollment", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def create_enrollment(data: EnrollmentCreateSchema, db: AsyncSession = Depends(get_db)):
    """
    Mendaftarkan mahasiswa ke student class (kelas mahasiswa).
    """
    # 1. Validasi mahasiswa exists
    mhs_stmt = select(Mahasiswa).where(Mahasiswa.nim == data.nim)
    mhs = (await db.execute(mhs_stmt)).scalars().first()
    if not mhs:
        raise HTTPException(404, f"Mahasiswa dengan NIM {data.nim} tidak ditemukan")
    
    # 2. Validasi student class exists
    sc_stmt = select(StudentClass).where(StudentClass.class_id == data.student_class_id)
    student_class = (await db.execute(sc_stmt)).scalars().first()
    if not student_class:
        raise HTTPException(404, f"Student Class dengan ID {data.student_class_id} tidak ditemukan")
    
    # 3. Check duplicate enrollment
    dup_stmt = select(KelasEnrollment).where(
        and_(
            KelasEnrollment.nim == data.nim,
            KelasEnrollment.student_class_id == data.student_class_id
        )
    )
    existing = (await db.execute(dup_stmt)).scalars().first()
    if existing:
        raise HTTPException(400, f"Mahasiswa {data.nim} sudah terdaftar di kelas {student_class.class_name}")
    
    # 4. Create enrollment
    new_enrollment = KelasEnrollment(
        nim=data.nim,
        student_class_id=data.student_class_id
    )
    db.add(new_enrollment)
    await db.commit()
    
    logger.info(f"Enrollment Created: {data.nim} -> Student Class {student_class.class_name}")
    return {"msg": f"Mahasiswa {data.nim} berhasil ditambahkan ke kelas {student_class.class_name}"}

@app.post("/admin/enrollment/bulk", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def bulk_create_enrollment(data: EnrollmentBulkSchema, db: AsyncSession = Depends(get_db)):
    """
    Bulk enrollment mahasiswa ke student class.
    """
    # 1. Validasi student class exists
    sc_stmt = select(StudentClass).where(StudentClass.class_id == data.kelas_id)
    student_class = (await db.execute(sc_stmt)).scalars().first()
    if not student_class:
        raise HTTPException(404, f"Student Class dengan ID {data.kelas_id} tidak ditemukan")
    
    success_count = 0
    failed_nims = []
    
    for nim in data.nim_list:
        try:
            # Check mahasiswa exists
            mhs_stmt = select(Mahasiswa).where(Mahasiswa.nim == nim)
            mhs = (await db.execute(mhs_stmt)).scalars().first()
            if not mhs:
                failed_nims.append(f"{nim} (tidak ditemukan)")
                continue
            
            # Check duplicate
            dup_stmt = select(KelasEnrollment).where(
                and_(
                    KelasEnrollment.nim == nim,
                    KelasEnrollment.student_class_id == data.kelas_id
                )
            )
            existing = (await db.execute(dup_stmt)).scalars().first()
            if existing:
                failed_nims.append(f"{nim} (sudah terdaftar)")
                continue
            
            # Create enrollment
            new_enrollment = KelasEnrollment(nim=nim, student_class_id=data.kelas_id)
            db.add(new_enrollment)
            success_count += 1
            
        except Exception as e:
            logger.error(f"Bulk Enrollment: Error enrolling {nim} - {e}")
            failed_nims.append(f"{nim} (error)")
    
    await db.commit()
    
    result_msg = f"Berhasil mendaftarkan {success_count} mahasiswa ke kelas {student_class.class_name}"
    if failed_nims:
        result_msg += f". Gagal: {', '.join(failed_nims)}"
    
    logger.info(f"Bulk Enrollment: {success_count} success, {len(failed_nims)} failed")
    return {"msg": result_msg, "success": success_count, "failed": len(failed_nims)}


@app.post("/admin/enrollment/bulk-excel", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def bulk_create_enrollment_excel(
    student_class_id: int = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk enrollment via Excel file upload to student class.
    Excel file harus memiliki kolom 'NIM' yang berisi list NIM mahasiswa.
    """
    # 1. Validasi student class exists
    sc_stmt = select(StudentClass).where(StudentClass.class_id == student_class_id)
    student_class = (await db.execute(sc_stmt)).scalars().first()
    if not student_class:
        raise HTTPException(404, f"Student Class dengan ID {student_class_id} tidak ditemukan")
    
    # 2. Validasi file extension
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "File harus berformat Excel (.xlsx atau .xls)")
    
    try:
        # 3. Read Excel file
        contents = await file.read()
        df = pd.read_excel(contents)
        
        # 4. Validate 'NIM' column exists
        if 'NIM' not in df.columns:
            raise HTTPException(400, "File Excel harus memiliki kolom 'NIM'")
        
        # 5. Extract NIMs and remove empty values
        nim_list = df['NIM'].dropna().astype(str).str.strip().tolist()
        
        if len(nim_list) == 0:
            raise HTTPException(400, "Tidak ada NIM yang ditemukan di file Excel")
        
        success_count = 0
        failed_nims = []
        
        # 6. Process each NIM
        for nim in nim_list:
            try:
                # Check mahasiswa exists
                mhs_stmt = select(Mahasiswa).where(Mahasiswa.nim == nim)
                mhs = (await db.execute(mhs_stmt)).scalars().first()
                if not mhs:
                    failed_nims.append(f"{nim} (tidak ditemukan)")
                    continue
                
                # Check duplicate
                dup_stmt = select(KelasEnrollment).where(
                    and_(
                        KelasEnrollment.nim == nim,
                        KelasEnrollment.student_class_id == student_class_id
                    )
                )
                existing = (await db.execute(dup_stmt)).scalars().first()
                if existing:
                    failed_nims.append(f"{nim} (sudah terdaftar)")
                    continue
                
                # Create enrollment
                new_enrollment = KelasEnrollment(nim=nim, student_class_id=student_class_id)
                db.add(new_enrollment)
                success_count += 1
                
            except Exception as e:
                logger.error(f"Bulk Enrollment Excel: Error enrolling {nim} - {e}")
                failed_nims.append(f"{nim} (error)")
        
        await db.commit()
        
        result_msg = f"Berhasil mendaftarkan {success_count} mahasiswa dari file Excel ke kelas {student_class.class_name}"
        if failed_nims:
            result_msg += f". Gagal: {', '.join(failed_nims[:10])}"
            if len(failed_nims) > 10:
                result_msg += f" ... dan {len(failed_nims) - 10} lainnya"
        
        logger.info(f"Bulk Enrollment Excel: {success_count} success, {len(failed_nims)} failed from file {file.filename}")
        return {"msg": result_msg, "success": success_count, "failed": len(failed_nims)}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk Enrollment Excel: Error reading file - {e}")
        raise HTTPException(500, f"Gagal memproses file Excel: {str(e)}")


@app.get("/admin/enrollment", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def get_all_enrollments(db: AsyncSession = Depends(get_db)):
    """
    Mendapatkan semua data enrollment dengan info mahasiswa dan kelas.
    """
    stmt = select(
        KelasEnrollment.enrollment_id,
        KelasEnrollment.nim,
        KelasEnrollment.student_class_id,
        KelasEnrollment.enrolled_at,
        Users.full_name.label('nama'),
        StudentClass.class_name
    ).join(
        Mahasiswa, KelasEnrollment.nim == Mahasiswa.nim
    ).join(
        Users, Mahasiswa.user_id == Users.user_id
    ).join(
        StudentClass, KelasEnrollment.student_class_id == StudentClass.class_id
    ).order_by(KelasEnrollment.enrollment_id.desc())
    
    result = await db.execute(stmt)
    enrollments = []
    
    for row in result.all():
        enrollments.append({
            "enrollment_id": row.enrollment_id,
            "nim": row.nim,
            "nama": row.nama or f"Mahasiswa {row.nim}",
            "student_class_id": row.student_class_id,
            "class_name": row.class_name,
            "enrolled_at": row.enrolled_at.isoformat() if row.enrolled_at else None
        })
    
    return enrollments

@app.get("/admin/enrollment/export-excel", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def export_enrollments_excel(db: AsyncSession = Depends(get_db)):
    """
    Export semua data enrollment ke file Excel.
    Menghasilkan file dengan kolom: NIM, Nama, Mata Kuliah, Kode Ruang, Tanggal Enroll.
    """
    try:
        # Get all enrollments with student and class info
        stmt = select(
            KelasEnrollment, 
            Users.full_name, 
            StudentClass.class_name
        ).join(
            Mahasiswa, KelasEnrollment.nim == Mahasiswa.nim
        ).outerjoin(
            Users, Mahasiswa.user_id == Users.user_id
        ).join(
            StudentClass, KelasEnrollment.student_class_id == StudentClass.class_id
        ).order_by(StudentClass.class_name, KelasEnrollment.nim)
        
        result = await db.execute(stmt)
        
        # Prepare data for Excel
        data = []
        for row in result.all():
            enrollment, nama, class_name = row
            data.append({
                'NIM': enrollment.nim,
                'Nama Mahasiswa': nama or f'Mahasiswa {enrollment.nim}',
                'Kelas Mahasiswa': class_name,
                'Tanggal Enroll': enrollment.enrolled_at.strftime('%Y-%m-%d %H:%M:%S') if enrollment.enrolled_at else ''
            })
        
        if len(data) == 0:
            raise HTTPException(404, "Tidak ada data enrollment untuk diekspor")
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        from io import BytesIO
        output = BytesIO()
        
        # Use xlsxwriter engine for better formatting
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Enrollment Data', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Enrollment Data']
            
            # Define formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4F46E5',
                'font_color': 'white',
                'border': 1
            })
            
            # Apply header format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Auto-fit columns
            for i, col in enumerate(df.columns):
                max_len = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                worksheet.set_column(i, i, max_len)
        
        output.seek(0)
        
        # Generate filename with timestamp
        from datetime import datetime
        filename = f"enrollment_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"Export Excel: Generated {filename} with {len(data)} records")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export Excel: Error - {e}")
        raise HTTPException(500, f"Gagal membuat file Excel: {str(e)}")

@app.get("/admin/enrollment/kelas/{student_class_id}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def get_enrollments_by_student_class(student_class_id: int, db: AsyncSession = Depends(get_db)):
    """
    Mendapatkan daftar mahasiswa yang terdaftar di student class tertentu.
    """
    stmt = select(
        KelasEnrollment.enrollment_id,
        KelasEnrollment.nim,
        KelasEnrollment.student_class_id,
        KelasEnrollment.enrolled_at,
        Users.full_name.label('nama'),
        StudentClass.class_name
    ).join(
        Mahasiswa, KelasEnrollment.nim == Mahasiswa.nim
    ).join(
        Users, Mahasiswa.user_id == Users.user_id
    ).join(
        StudentClass, KelasEnrollment.student_class_id == StudentClass.class_id
    ).where(
        KelasEnrollment.student_class_id == student_class_id
    ).order_by(KelasEnrollment.nim)
    
    result = await db.execute(stmt)
    enrollments = []
    
    for row in result.all():
        enrollments.append({
            "enrollment_id": row.enrollment_id,
            "nim": row.nim,
            "nama": row.nama or f"Mahasiswa {row.nim}",
            "student_class_id": row.student_class_id,
            "class_name": row.class_name,
            "enrolled_at": row.enrolled_at.isoformat() if row.enrolled_at else None
        })
    
    return enrollments

# ==============================================================================
# [SECTION 11.6] STUDENT CLASS MANAGEMENT (Kelas Mahasiswa: A11.4109, etc)
# ==============================================================================

@app.get("/admin/student-class", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def get_all_student_classes(db: AsyncSession = Depends(get_db)):
    """Get all student classes with enrollment count."""
    stmt = select(
        StudentClass,
        func.count(KelasEnrollment.enrollment_id).label('student_count')
    ).outerjoin(
        KelasEnrollment, StudentClass.class_id == KelasEnrollment.student_class_id
    ).group_by(StudentClass.class_id).order_by(StudentClass.class_name)
    
    result = await db.execute(stmt)
    classes = []
    for row in result.all():
        student_class, count = row
        classes.append({
            "class_id": student_class.class_id,
            "class_name": student_class.class_name,
            "student_count": count,
            "created_at": student_class.created_at.isoformat() if student_class.created_at else None
        })
    return classes

@app.post("/admin/student-class", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def create_student_class(class_name: str = Form(...), db: AsyncSession = Depends(get_db)):
    """Create new student class."""
    dup = (await db.execute(select(StudentClass).where(StudentClass.class_name == class_name))).scalars().first()
    if dup:
        raise HTTPException(400, f"Kelas {class_name} sudah ada")
    new_class = StudentClass(class_name=class_name)
    db.add(new_class)
    await db.commit()
    await db.refresh(new_class)
    logger.info(f"Student Class Created: {class_name}")
    return {"msg": f"Kelas {class_name} berhasil dibuat", "class_id": new_class.class_id}

@app.get("/admin/student-class/{class_id}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def get_student_class_detail(class_id: int, db: AsyncSession = Depends(get_db)):
    """Get student class detail with enrolled students."""
    sc = (await db.execute(select(StudentClass).where(StudentClass.class_id == class_id))).scalars().first()
    if not sc:
        raise HTTPException(404, "Student Class tidak ditemukan")
    
    # Get enrolled students
    stmt = select(
        KelasEnrollment.enrollment_id,
        KelasEnrollment.nim,
        Users.full_name
    ).join(
        Mahasiswa, KelasEnrollment.nim == Mahasiswa.nim
    ).join(
        Users, Mahasiswa.user_id == Users.user_id
    ).where(
        KelasEnrollment.student_class_id == class_id
    ).order_by(KelasEnrollment.nim)
    
    result = await db.execute(stmt)
    students = []
    for row in result.all():
        students.append({
            "enrollment_id": row.enrollment_id,
            "nim": row.nim,
            "nama": row.full_name or f"Mahasiswa {row.nim}"
        })
    
    return {
        "class_id": sc.class_id,
        "class_name": sc.class_name,
        "created_at": sc.created_at.isoformat() if sc.created_at else None,
        "students": students
    }

@app.post("/admin/mahasiswa/bulk-face-registration", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def bulk_face_registration(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk Face Registration - Upload ZIP berisi foto mahasiswa.
    
    **Workflow:**
    1. Upload ZIP file containing student photos
    2. System extracts NIM from filename (e.g., A11.2025.16442.jpg)
    3. Detect face in each image (must be exactly 1 face)
    4. Generate face embeddings using InsightFace
    5. Save/Update to database
    6. Return detailed summary report
    
    **File Naming Convention:**
    - Format: {NIM}.jpg or {NIM}.png
    - Example: A11.2025.16442.jpg
    - System will extract 'A11.2025.16442' as NIM
    
    **Validation Rules:**
    - ✅ Exactly 1 face detected → Success
    - ❌ 0 faces detected → Failed (No face found)
    - ❌ Multiple faces detected → Failed (Use single face photo)
    
    **Response:**
    - Total processed
    - Success count
    - Failed count
    - Detailed list of results with error reasons
    """
    from bulk_face_registration import (
        BulkFaceProcessor,
        validate_zip_file,
        is_valid_zip_file
    )
    
    # Validate file type
    if not file.filename.lower().endswith('.zip'):
        raise HTTPException(400, "Only ZIP files are allowed")
    
    # Validate file size
    file_size = 0
    content = await file.read()
    file_size = len(content)
    await file.seek(0)  # Reset file pointer
    
    is_valid, error_msg = validate_zip_file(file_size, max_size_mb=500)
    if not is_valid:
        raise HTTPException(400, error_msg)
    
    # Save ZIP temporarily
    temp_zip_path = f"temp_videos/bulk_face_{uuid.uuid4().hex}.zip"
    os.makedirs("temp_videos", exist_ok=True)
    
    try:
        # Write uploaded file to disk
        async with aiofiles.open(temp_zip_path, 'wb') as f:
            await f.write(content)
        
        # Validate ZIP integrity
        is_valid_zip, zip_error = is_valid_zip_file(temp_zip_path)
        if not is_valid_zip:
            raise HTTPException(400, f"Invalid ZIP file: {zip_error}")
        
        logger.info(f"📦 Bulk Face Registration started - File: {file.filename} ({file_size / 1024 / 1024:.2f} MB)")
        
        # Initialize processor
        processor = BulkFaceProcessor(db)
        
        # Process ZIP file
        summary = await processor.process_zip_file(temp_zip_path)
        
        logger.info(f"✅ Bulk registration completed - Success: {summary.success_count}/{summary.total_processed}")
        
        return summary
    
    except Exception as e:
        logger.error(f"❌ Bulk face registration error: {str(e)}")
        raise HTTPException(500, f"Processing error: {str(e)}")
    
    finally:
        # Cleanup ZIP file
        try:
            if os.path.exists(temp_zip_path):
                os.remove(temp_zip_path)
        except Exception as e:
            logger.warning(f"Failed to cleanup ZIP file: {e}")

@app.delete("/admin/student-class/{class_id}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def delete_student_class(class_id: int, force: bool = False, db: AsyncSession = Depends(get_db)):
    """Delete student class. Use force=true to delete with enrollments and schedules."""
    sc = (await db.execute(select(StudentClass).where(StudentClass.class_id == class_id))).scalars().first()
    if not sc:
        raise HTTPException(404, "Student Class tidak ditemukan")
    
    # Count related records
    enr_count = (await db.execute(select(func.count()).select_from(KelasEnrollment).where(KelasEnrollment.student_class_id == class_id))).scalar()
    jadwal_count = (await db.execute(select(func.count()).select_from(Jadwal).where(Jadwal.student_class_id == class_id))).scalar()
    
    # If force=true, delete all related records first
    if force:
        # Delete jadwal first
        if jadwal_count > 0:
            await db.execute(delete(Jadwal).where(Jadwal.student_class_id == class_id))
            logger.info(f"Deleted {jadwal_count} jadwal from class {sc.class_name}")
        
        # Delete enrollments
        if enr_count > 0:
            await db.execute(delete(KelasEnrollment).where(KelasEnrollment.student_class_id == class_id))
            logger.info(f"Deleted {enr_count} enrollments from class {sc.class_name}")
    elif enr_count > 0 or jadwal_count > 0:
        raise HTTPException(400, f"Tidak bisa menghapus kelas yang masih memiliki {enr_count} mahasiswa dan {jadwal_count} jadwal. Gunakan force=true untuk hapus paksa.")
    
    await db.delete(sc)
    await db.commit()
    logger.info(f"Student Class Deleted: {sc.class_name}")
    return {
        "msg": f"Kelas {sc.class_name} berhasil dihapus", 
        "enrollments_deleted": enr_count if force else 0,
        "jadwal_deleted": jadwal_count if force else 0
    }

@app.put("/admin/student-class/{class_id}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def update_student_class(class_id: int, class_name: str = Form(...), db: AsyncSession = Depends(get_db)):
    """Update/rename student class."""
    sc = (await db.execute(select(StudentClass).where(StudentClass.class_id == class_id))).scalars().first()
    if not sc:
        raise HTTPException(404, "Student Class tidak ditemukan")
    
    # Check duplicate name
    dup = (await db.execute(select(StudentClass).where(
        and_(StudentClass.class_name == class_name, StudentClass.class_id != class_id)
    ))).scalars().first()
    if dup:
        raise HTTPException(400, f"Kelas {class_name} sudah ada")
    
    old_name = sc.class_name
    sc.class_name = class_name
    await db.commit()
    logger.info(f"Student Class Renamed: {old_name} -> {class_name}")
    return {"msg": f"Kelas berhasil diubah dari {old_name} menjadi {class_name}"}

@app.get("/admin/enrollment/mahasiswa/{nim}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def get_enrollments_by_mahasiswa(nim: str, db: AsyncSession = Depends(get_db)):
    """
    Mendapatkan daftar student class yang diikuti oleh mahasiswa tertentu.
    """
    stmt = select(
        KelasEnrollment,
        StudentClass.class_name
    ).join(
        StudentClass, KelasEnrollment.student_class_id == StudentClass.class_id
    ).where(
        KelasEnrollment.nim == nim
    ).order_by(StudentClass.class_name)
    
    result = await db.execute(stmt)
    classes = []
    
    for row in result.all():
        enrollment, class_name = row
        classes.append({
            "enrollment_id": enrollment.enrollment_id,
            "student_class_id": enrollment.student_class_id,
            "class_name": class_name,
            "enrolled_at": enrollment.enrolled_at.isoformat() if enrollment.enrolled_at else None
        })
    
    return classes

@app.delete("/admin/enrollment/{enrollment_id}", dependencies=[Depends(allow_kaprodi)], tags=["Admin"])
async def delete_enrollment(enrollment_id: int, db: AsyncSession = Depends(get_db)):
    """
    Menghapus enrollment mahasiswa dari kelas.
    """
    stmt = select(KelasEnrollment).where(KelasEnrollment.enrollment_id == enrollment_id)
    enrollment = (await db.execute(stmt)).scalars().first()
    
    if not enrollment:
        raise HTTPException(404, "Enrollment tidak ditemukan")
    
    nim = enrollment.nim
    kelas_id = enrollment.kelas_id
    
    await db.delete(enrollment)
    await db.commit()
    
    logger.info(f"Enrollment: Removed {nim} from kelas_id {kelas_id}")
    return {"msg": "Enrollment berhasil dihapus"}

# ==============================================================================
# [SECTION 12] MAHASISWA & REPORTING
# ==============================================================================

@app.get("/mhs/jadwal", dependencies=[Depends(allow_mhs)], tags=["Mahasiswa"])
async def student_get_jadwal(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """
    [UPDATED] Menampilkan jadwal kuliah yang di-enroll oleh mahasiswa.
    Mahasiswa hanya melihat jadwal untuk kelas yang mereka ikuti.
    Includes jadwal_id untuk tracking status kehadiran per jadwal.
    """
    nim = user['username']
    
    # Query dengan JOIN ke KelasEnrollment untuk filter berdasarkan enrollment
    stmt = select(
        Jadwal, 
        Users.full_name, 
        Kelas.nama_matkul, 
        Kelas.kode_ruang,
        StudentClass.class_name
    ).join(
        Users, Jadwal.dosen_username == Users.username
    ).join(
        Kelas, Jadwal.kelas_id == Kelas.kelas_id
    ).outerjoin(
        StudentClass, Jadwal.student_class_id == StudentClass.class_id
    ).join(
        KelasEnrollment, 
        and_(
            KelasEnrollment.student_class_id == Jadwal.student_class_id,
            KelasEnrollment.nim == nim
        )
    )
    
    result = await db.execute(stmt)
    data = []
    for row in result.all():
        data.append({
            "jadwal_id": row[0].jadwal_id,
            "hari": row[0].hari,
            "jam": f"{row[0].jam_mulai}-{row[0].jam_selesai}",
            "jam_mulai": row[0].jam_mulai,
            "jam_selesai": row[0].jam_selesai,
            "dosen": row[1],
            "matkul": f"{row[2]} ({row[3]})",
            "ruang": row[3],
            "waktu": f"{row[0].hari}, {row[0].jam_mulai}",
            "student_class_name": row[4]
        })
    return data

@app.get("/mhs/history", dependencies=[Depends(allow_mhs)], tags=["Mahasiswa"])
async def student_get_history(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """
    Melihat riwayat absensi diri sendiri dengan detail mata kuliah.
    Join dengan Jadwal dan Kelas untuk mendapatkan nama matkul, ruang, dll.
    Hanya menampilkan log dari kelas yang di-enroll.
    """
    nim = user['username']
    
    # Query dengan join untuk mendapatkan info lengkap
    # Tambahkan join dengan KelasEnrollment untuk memastikan hanya log dari kelas enrolled
    stmt = select(
        LogAbsensi,
        Jadwal.hari,
        Jadwal.jam_mulai,
        Jadwal.jam_selesai,
        Kelas.nama_matkul,
        Kelas.kode_ruang
    ).join(
        Jadwal, LogAbsensi.jadwal_id == Jadwal.jadwal_id
    ).join(
        Kelas, Jadwal.kelas_id == Kelas.kelas_id
    ).join(
        KelasEnrollment,
        and_(
            KelasEnrollment.student_class_id == Jadwal.student_class_id,
            KelasEnrollment.nim == nim
        )
    ).where(
        LogAbsensi.nim == nim
    ).order_by(desc(LogAbsensi.waktu_absen))
    
    result = await db.execute(stmt)
    rows = result.all()
    
    # Format response dengan info lengkap
    data = []
    for row in rows:
        log = row[0]
        data.append({
            "log_id": log.log_id,
            "jadwal_id": log.jadwal_id,
            "waktu_absen": log.waktu_absen.isoformat() if log.waktu_absen else None,
            "metode": log.metode,
            "is_disputed": log.is_disputed,
            "keterangan_report": log.keterangan_report,
            "bukti_foto": log.bukti_foto,
            "jumlah_muncul": log.jumlah_muncul,
            "emosi_dominan": log.emosi_dominan,
            # Info tambahan dari Jadwal & Kelas
            "hari": row[1],
            "jam": f"{row[2]}-{row[3]}" if row[2] and row[3] else None,
            "matkul": row[4],  # nama_matkul
            "ruang": row[5]    # kode_ruang
        })
    
    return data

@app.post("/mhs/report", dependencies=[Depends(allow_mhs)], tags=["Mahasiswa"])
async def student_submit_report(
    log_id: int = Form(...), 
    alasan: str = Form(...), 
    foto: Optional[UploadFile] = File(None),
    user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Mengajukan sengketa kehadiran dengan bukti foto opsional.
    Foto akan disimpan di folder hasil_crop/reports/ untuk ditinjau dosen.
    """
    # Cek apakah log absensi ada dan milik mahasiswa ini
    log_stmt = select(LogAbsensi).options(joinedload(LogAbsensi.jadwal)).where(LogAbsensi.log_id == log_id)
    existing_log = (await db.execute(log_stmt)).scalars().first()
    
    if not existing_log:
        raise HTTPException(status_code=404, detail="Log absensi tidak ditemukan")
    
    if existing_log.nim != user['username']:
        raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke log ini")
    
    # Validasi enrollment: pastikan mahasiswa enrolled di kelas yang sesuai
    if existing_log.jadwal and existing_log.jadwal.student_class_id:
        enrollment_stmt = select(KelasEnrollment).where(
            and_(
                KelasEnrollment.nim == user['username'],
                KelasEnrollment.student_class_id == existing_log.jadwal.student_class_id
            )
        )
        enrollment = (await db.execute(enrollment_stmt)).scalars().first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Anda tidak terdaftar di kelas ini")
    
    # Simpan foto bukti jika ada
    bukti_path = existing_log.bukti_foto  # Pertahankan foto lama jika ada
    
    if foto:
        try:
            # Buat folder reports jika belum ada
            reports_dir = os.path.join(DIRECTORY_CONFIG["CROP_FACE"], "reports")
            os.makedirs(reports_dir, exist_ok=True)
            
            # Generate nama file unik
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = os.path.splitext(foto.filename)[1] or '.jpg'
            filename = f"report_{user['username']}_{log_id}_{timestamp}{ext}"
            filepath = os.path.join(reports_dir, filename)
            
            # Simpan file
            async with aiofiles.open(filepath, 'wb') as f:
                content = await foto.read()
                await f.write(content)
            
            bukti_path = f"/hasil_crop/reports/{filename}"
            logger.info(f"Report: Foto bukti disimpan di {bukti_path}")
            
        except Exception as e:
            logger.error(f"Report: Gagal menyimpan foto bukti - {e}")
            # Lanjutkan tanpa foto jika gagal
    
    # Update log absensi
    await db.execute(
        update(LogAbsensi)
        .where(LogAbsensi.log_id == log_id)
        .values(
            is_disputed=True, 
            keterangan_report=alasan,
            bukti_foto=bukti_path
        )
    )
    await db.commit()
    
    logger.info(f"Report: Mahasiswa {user['username']} melaporkan log_id {log_id}")
    return {"msg": "Laporan berhasil dikirim", "bukti_foto": bukti_path}

@app.get("/download_excel/{task_id}", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def download_attendance_excel(task_id: str, db: AsyncSession = Depends(get_db)):
    """Generate Excel Report per Sesi Video (AI detected only)."""
    # Ambil jadwal_id dari task
    task_stmt = select(VideoTask).where(VideoTask.task_id == task_id)
    task = (await db.execute(task_stmt)).scalars().first()
    
    if task and task.jadwal_id:
        # Jika ada jadwal_id, ambil SEMUA absensi untuk jadwal ini hari ini
        today = datetime.now().date()
        stmt = select(LogAbsensi, Users.full_name).outerjoin(
            Mahasiswa, LogAbsensi.nim == Mahasiswa.nim
        ).outerjoin(
            Users, Mahasiswa.user_id == Users.user_id
        ).where(
            and_(
                LogAbsensi.jadwal_id == task.jadwal_id,
                func.date(LogAbsensi.waktu_absen) == today
            )
        ).order_by(LogAbsensi.nim)
    else:
        # Fallback: filter by task_id saja
        stmt = select(LogAbsensi, Users.full_name).outerjoin(
            Mahasiswa, LogAbsensi.nim == Mahasiswa.nim
        ).outerjoin(
            Users, Mahasiswa.user_id == Users.user_id
        ).where(LogAbsensi.task_id == task_id).order_by(LogAbsensi.nim)
    
    rows = (await db.execute(stmt)).all()
    
    if not rows: return HTMLResponse("Belum ada data untuk sesi ini.", 404)
    
    # Convert timezone-aware datetime to timezone-naive string for Excel compatibility
    data = []
    for r in rows:
        waktu = r[0].waktu_absen
        # Format datetime sebagai string agar Excel tidak error
        if waktu:
            waktu_str = waktu.strftime("%Y-%m-%d %H:%M:%S") if hasattr(waktu, 'strftime') else str(waktu)
        else:
            waktu_str = "-"
        
        data.append({
            "NIM": r[0].nim, 
            "Nama": r[1] or "-",
            "Waktu": waktu_str, 
            "Metode": r[0].metode or "-",
            "Status": "SENGKETA" if r[0].is_disputed else "HADIR"
        })
    
    df = pd.DataFrame(data)
    
    path = os.path.join(DIRECTORY_CONFIG["EXCEL_REPORT"], f"{task_id}.xlsx")
    df.to_excel(path, index=False)
    
    return FileResponse(path, filename=f"Laporan_Absensi_{task_id}.xlsx")

# ==============================================================================
# [SECTION 13] UTILITY (REGISTER & USERS)
# ==============================================================================

@app.post("/register/", dependencies=[Depends(allow_dosen)], tags=["Dosen"])
async def register_face_endpoint(
    nim: str = Form(...), 
    password: str = Form(...),
    file: UploadFile = File(...), 
    db: AsyncSession = Depends(get_db)
):
    """
    Registrasi Wajah Mahasiswa dengan Password.
    [AUTO-LINK] Otomatis membuat akun User jika belum ada untuk mencegah error relasi.
    Password akan di-hash dengan bcrypt sebelum disimpan.
    """
    if state.face_app is None: raise HTTPException(503, "AI belum siap")
    
    content = await file.read()
    img = cv2.imdecode(np.frombuffer(content, np.uint8), cv2.IMREAD_COLOR)
    faces = state.face_app.get(img)
    if not faces: raise HTTPException(400, "Wajah tidak terdeteksi")
    
    emb = faces[0].embedding / np.linalg.norm(faces[0].embedding)
    
    # 1. Cek User Account
    user_q = await db.execute(select(Users).where(Users.username == nim))
    user_obj = user_q.scalars().first()
    
    if not user_obj:
        logger.info(f"Register: Creating User for {nim} with custom password")
        # Hash password dengan bcrypt
        hashed_password = get_password_hash(password)
        user_obj = Users(
            username=nim, 
            password=hashed_password, 
            full_name=f"Mahasiswa {nim}", 
            role="mahasiswa"
        )
        db.add(user_obj)
        await db.flush()
    else:
        # Update password jika user sudah ada
        logger.info(f"Register: Updating password for existing user {nim}")
        user_obj.password = get_password_hash(password)
    
    # 2. Upsert Mahasiswa
    check = await db.execute(select(Mahasiswa).where(Mahasiswa.nim == nim))
    existing = check.scalars().first()
    
    if existing:
        existing.embedding_data = emb.tolist()
        existing.user_id = user_obj.user_id # Ensure Link
    else:
        db.add(Mahasiswa(nim=nim, user_id=user_obj.user_id, embedding_data=emb.tolist()))
        
    await db.commit()
    await reload_face_database_async()
    return {"status": "success", "message": f"Wajah {nim} berhasil didaftarkan dengan password ter-enkripsi"}

@app.get("/users/", dependencies=[Depends(allow_all)], tags=["Common"])
async def list_all_users(db: AsyncSession = Depends(get_db)):
    """List semua mahasiswa untuk dropdown."""
    stmt = select(Mahasiswa).options(joinedload(Mahasiswa.user))
    res = (await db.execute(stmt)).scalars().all()
    return [{"nim": s.nim, "nama": s.user.full_name if s.user else s.nim} for s in res]

# ==============================================================================
# [SECTION 13.5] ADMIN CRUD - KELAS & JADWAL
# ==============================================================================

class EnrollmentBulkSchema(BaseModel):
    kelas_id: int
    nim_list: List[str]

# Student Class Management Schemas
class StudentClassCreateSchema(BaseModel):
    class_name: str  # A11.4109

class StudentClassUpdateSchema(BaseModel):
    class_name: str

class DisputeReportSchema(BaseModel):
    dosen_username: str
    kelas_id: int
    hari: str
    jam_mulai: str
    jam_selesai: str

class KelasUpdateSchema(BaseModel):
    nama_matkul: str
    kode_ruang: str

class JadwalUpdateSchema(BaseModel):
    dosen_username: str
    kelas_id: int
    student_class_id: Optional[int] = None  # Optional: Student Class
    hari: str
    jam_mulai: str
    jam_selesai: str

@app.put("/admin/kelas/{kelas_id}", dependencies=[Depends(allow_all)], tags=["Admin"])
async def update_kelas(kelas_id: int, data: KelasUpdateSchema, db: AsyncSession = Depends(get_db)):
    """Update mata kuliah."""
    stmt = select(Kelas).where(Kelas.kelas_id == kelas_id)
    result = (await db.execute(stmt)).scalars().first()
    if not result:
        raise HTTPException(404, "Mata kuliah tidak ditemukan")
    
    result.nama_matkul = data.nama_matkul
    result.kode_ruang = data.kode_ruang
    await db.commit()
    return {"message": "Mata kuliah berhasil diperbarui"}

@app.delete("/admin/kelas/{kelas_id}", dependencies=[Depends(allow_all)], tags=["Admin"])
async def delete_kelas(kelas_id: int, db: AsyncSession = Depends(get_db)):
    """Hapus mata kuliah."""
    stmt = select(Kelas).where(Kelas.kelas_id == kelas_id)
    result = (await db.execute(stmt)).scalars().first()
    if not result:
        raise HTTPException(404, "Mata kuliah tidak ditemukan")
    
    await db.delete(result)
    await db.commit()
    return {"message": "Mata kuliah berhasil dihapus"}

@app.put("/admin/jadwal/{jadwal_id}", dependencies=[Depends(allow_all)], tags=["Admin"])
async def update_jadwal(jadwal_id: int, data: JadwalUpdateSchema, db: AsyncSession = Depends(get_db)):
    """Update jadwal kuliah."""
    stmt = select(Jadwal).where(Jadwal.jadwal_id == jadwal_id)
    result = (await db.execute(stmt)).scalars().first()
    if not result:
        raise HTTPException(404, "Jadwal tidak ditemukan")
    
    result.dosen_username = data.dosen_username
    result.kelas_id = data.kelas_id
    result.student_class_id = data.student_class_id  # NEW: Student Class
    result.hari = data.hari
    result.jam_mulai = data.jam_mulai
    result.jam_selesai = data.jam_selesai
    await db.commit()
    return {"message": "Jadwal berhasil diperbarui"}

@app.delete("/admin/jadwal/{jadwal_id}", dependencies=[Depends(allow_all)], tags=["Admin"])
async def delete_jadwal(jadwal_id: int, db: AsyncSession = Depends(get_db)):
    """Hapus jadwal kuliah."""
    stmt = select(Jadwal).where(Jadwal.jadwal_id == jadwal_id)
    result = (await db.execute(stmt)).scalars().first()
    if not result:
        raise HTTPException(404, "Jadwal tidak ditemukan")
    
    await db.delete(result)
    await db.commit()
    return {"message": "Jadwal berhasil dihapus"}

# ==============================================================================
# [SECTION 14] FRONTEND ROUTING (SPA SUPPORT)
# ==============================================================================
@app.get("/login.html", response_class=HTMLResponse)
async def r_login(): return FileResponse(os.path.join(BASE_DIR, "login.html"))

@app.get("/register-page", response_class=HTMLResponse)
async def r_reg(): return FileResponse(os.path.join(BASE_DIR, "register.html"))

@app.get("/dashboard/kaprodi", response_class=HTMLResponse)
async def r_kaprodi(): return FileResponse(os.path.join(BASE_DIR, "dashboard_kaprodi.html"))

@app.get("/dashboard/dosen", response_class=HTMLResponse)
async def r_dosen(): return FileResponse(os.path.join(BASE_DIR, "dashboard_dosen.html"))

@app.get("/dashboard/mahasiswa", response_class=HTMLResponse)
async def r_mhs(): return FileResponse(os.path.join(BASE_DIR, "dashboard_mahasiswa.html"))

@app.get("/")
async def root(): return RedirectResponse("/login.html")